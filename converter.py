import os

import cv2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from card import Card
from openpyxl.styles import PatternFill

XTRACK_VS_FREQ = {'0110': 12,
                '0101': 8,
                '1010': 8,
                '1001': 7,
                '0001': 3,
                '1000': 3,
                '0100': 1,
                '0010': 1}


def read_colors(colors_path):
    with open(colors_path) as file:
        colors_hex = file.readlines()
        colors_hex = [line.rstrip().lstrip('#') for line in colors_hex]
    
    return colors_hex


def read_card(img_path, target_colors):
    card = cv2.imread(img_path, cv2.COLOR_BGR2RGB)

    height = card.shape[0]
    colors_id = [i * height // 5 for i in range(1, 5)]

    half_width = card.shape[1] // 2
    colors = [card[i, half_width] for i in colors_id]
    
    card_obj = Card(colors, target_colors)
    
    return card_obj


target_colors = read_colors('colors.txt')
#cards_dirs = ['/home/egor/buhf6/balance/converter/prototypes/random_deck']
#cards_dirs = ['../../cards/base', '../../ca
cards_dirs = ['../../cards/test_22_08/']
#cards_dirs = ['../../cards/base', '../../cards/rotor', '../../cards/snatch']

cards_pathes = []
for cdir in cards_dirs:
    for file in os.listdir(cdir):
        filename = os.fsdecode(file)
        if filename.endswith(".png"): 
            path = os.path.join(cdir, filename)
            cards_pathes.append(path)
        else:
            continue

print(f"Reading {len(cards_pathes)} cards...")

cards_list = [read_card(path, target_colors) for path in cards_pathes]

tracks_vs_colors = {track: {color: 0 for color in card.color_tracks.keys()} 
                    for card in cards_list
                    for track in card.color_tracks.values()
                    }

for card in cards_list:
    color_tracks = card.color_tracks
    for color, track in color_tracks.items():
        try:
            tracks_vs_colors[track][color] += 1
        except:
            tracks_vs_colors[track]
            
result = pd.DataFrame.from_dict(tracks_vs_colors)
result = result.T
result = result.reindex(['0110', '0101', '1010', '1001', 
                         '0001', '1000', '0100', '0100', '0000'])


result = result.drop("0000")

good_freq_list = [freq for freq in XTRACK_VS_FREQ.values()]
good_freq_list = [freq / sum(good_freq_list) for freq in good_freq_list]
real_freq_list = result.sum(axis=1).to_numpy()
real_freq_list = [freq / sum(real_freq_list) for freq in real_freq_list]
x_axis = np.arange(len(real_freq_list))

bar_width = .3
labels = ['0110', '0101', '1010', '1001', 
          '0001', '1000', '0100', '0100']
plt.bar(x_axis, good_freq_list, width=bar_width, label='target')
plt.bar(x_axis + bar_width, real_freq_list, width=bar_width, label='real')
plt.xticks(ticks=x_axis, labels=labels)
plt.legend()
plt.show()

result['Tracks total'] = result.sum(axis=1)

result.to_excel('tracks_vs_colors.xlsx')

wb = openpyxl.load_workbook("tracks_vs_colors.xlsx") #path to the Excel file
ws = wb['Sheet1'] #Name of the working sheet


col_id = ['B', 'C', 'D', 'E']
row_id = 1

for col_num, col_letter in enumerate(col_id):
    cell_val = ws[f"{col_letter}{row_id}"].value
    print(cell_val)
    fill_cell = PatternFill(patternType='solid', fgColor=cell_val) 
    ws[f"{col_letter}{row_id}"].fill = fill_cell
wb.save("tracks_vs_colors.xlsx")



